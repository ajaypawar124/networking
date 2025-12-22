import paramiko
from lxml import etree
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment

USERNAME = "regress"
PASSWORD = "MaRtInI"
START_DEVICE = "10.13.124.163"
OUTPUT_FILE = "lldp_topology.xlsx"

visited = set()
links = set()

# --- Excel setup ---
wb = Workbook()

# Sheet 1 — LLDP links
ws_topo = wb.active
ws_topo.title = "LLDP Topology"
ws_topo.append([
    "Device Hostname", "Device IP", "Local Interface",
    "Neighbor Hostname", "Neighbor IP", "Neighbor Interface"
])

# Sheet 2 — Device summary (SSH status, serial, hardware)
ws_summary = wb.create_sheet("Device Summary")
ws_summary.append([
    "Timestamp", "Hostname", "IP", "SSH Status", "Serial Number", "Hardware Description"
])

# Sheet 3 — Disabled interfaces output
ws_disabled = wb.create_sheet("Disabled Interfaces")
ws_disabled.append(["Device Hostname", "Device IP", "Disabled Interfaces Output"])


def save_excel():
    wb.save(OUTPUT_FILE)
    print(f"[INFO] Excel saved to {OUTPUT_FILE}")


# --- SSH helper ---
def ssh_run(host, cmd, log_status=False):
    """Run command via SSH and return output. Optionally log SSH success/failure."""
    try:
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(
            host,
            username=USERNAME,
            password=PASSWORD,
            look_for_keys=False,
            allow_agent=False,
            timeout=10,
        )
        stdin, stdout, stderr = client.exec_command(cmd)
        out = stdout.read().decode(errors="ignore")
        client.close()

        if log_status:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws_summary.append([timestamp, "", host, "✅ Success", "", ""])
        return out
    except Exception as e:
        print(f"[ERROR] SSH to {host} failed: {e}")
        if log_status:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws_summary.append([timestamp, "", host, f"❌ Failed ({e})", "", ""])
        return ""


def get_hostname_and_disabled(ip):
    """Fetch hostname, disabled interfaces output, and hardware info."""
    xml = ssh_run(ip, "show version | display xml", log_status=True)
    try:
        root = etree.fromstring(xml.encode())
        hostname = root.xpath("//*[local-name()='host-name']/text()")
        name = hostname[0].strip() if hostname else ip
    except Exception:
        name = ip

    # --- Get Chassis hardware info ---
    chassis_xml = ssh_run(ip, "show chassis hardware | display xml", log_status=False)
    serial = ""
    hardware = ""
    try:
        root = etree.fromstring(chassis_xml.encode())
        serial = root.xpath("//*[local-name()='chassis']/*[local-name()='serial-number']/text()")
        hardware = root.xpath("//*[local-name()='chassis']/*[local-name()='description']/text()")
        serial = serial[0] if serial else ""
        hardware = hardware[0] if hardware else ""
    except Exception as e:
        print(f"[WARN] Failed to parse chassis info for {ip}: {e}")

    # --- Update Device Summary with hardware details ---
    for row in ws_summary.iter_rows(min_row=2, values_only=False):
        if row[2].value == ip:
            row[1].value = name
            row[4].value = serial
            row[5].value = hardware
            break

    # --- Collect disabled interfaces output ---
    disable_output = ssh_run(
        ip,
        "show configuration interfaces | display set | display inheritance | match disable",
        log_status=False,
    )
    if not disable_output:
        disable_output = "(No disabled interfaces found or device not reachable)"

    ws_disabled.append([name, ip, None])
    row = ws_disabled.max_row
    cell = ws_disabled.cell(row=row, column=3)
    cell.value = disable_output
    cell.alignment = Alignment(wrap_text=True)

    return name


def parse_lldp(xml_data):
    """Parse LLDP neighbor details using XPath (handles namespaces automatically)."""
    neighbors = []
    try:
        root = etree.fromstring(xml_data.encode())
    except Exception as e:
        print(f"[WARN] XML parse error: {e}")
        return neighbors

    neighbor_elems = root.xpath("//*[local-name()='lldp-neighbor-information']")

    for n in neighbor_elems:
        local_if = n.xpath("string(*[local-name()='lldp-local-interface'])") or "N/A"
        remote_name = n.xpath("string(*[local-name()='lldp-remote-system-name'])") or "N/A"
        remote_port = n.xpath("string(*[local-name()='lldp-remote-port-id'])") or "N/A"

        mgmt_ip = "N/A"
        addr_list = n.xpath(".//*[local-name()='lldp-remote-management-address']/text()")
        if addr_list:
            for a in addr_list:
                if "." in a:
                    mgmt_ip = a
                    break
            else:
                mgmt_ip = addr_list[0]

        neighbors.append((local_if, remote_name, mgmt_ip, remote_port))
    return neighbors


def explore(ip):
    """Recursive LLDP exploration."""
    if ip in visited:
        return
    visited.add(ip)

    hostname = get_hostname_and_disabled(ip)
    print(f"\n[INFO] Exploring {hostname} ({ip})")

    xml = ssh_run(ip, "show lldp neighbors detail | display xml", log_status=False)
    if not xml.strip():
        print(f"[WARN] No LLDP XML from {ip}")
        return

    neighs = parse_lldp(xml)
    print(f"[DEBUG] Found {len(neighs)} neighbors on {hostname}")

    for local_if, n_name, n_ip, n_port in neighs:
        # Create order-independent key to prevent duplicate links
        key = tuple(sorted([
            (ip, local_if),
            (n_ip, n_port)
        ]))
        if key in links:
            continue
        links.add(key)

        ws_topo.append([hostname, ip, local_if, n_name, n_ip, n_port])

        if n_ip == "N/A":
            print(f"  [WARN] No management IP for {n_name} ({local_if})")
        else:
            print(f"  [LINK] {hostname}:{local_if} ↔ {n_name}:{n_port} ({n_ip})")

    # Recurse for new devices
    for _, _, n_ip, _ in neighs:
        if n_ip not in visited and n_ip != "N/A" and "." in n_ip:
            time.sleep(1)
            explore(n_ip)


if __name__ == "__main__":
    explore(START_DEVICE)
    save_excel()
    print("\n✅ Discovery complete. Check 'lldp_topology.xlsx'.")

